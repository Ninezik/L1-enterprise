from fastapi import FastAPI, Query
from fastapi.responses import Response
import pandas as pd
import psycopg2
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime, timedelta

app = FastAPI()

# =========================
# KONFIG REDSHIFT
# =========================
DB_HOST = "pos-redshift.cwig526q7i0q.ap-southeast-3.redshift.amazonaws.com"
DB_PORT = "5439"
DB_NAME = "posind_kurlog"
DB_USER = "rda_analis"
DB_PASSWORD = "GcTz69eZ6UwNnRhypjx9Ysk8"


@app.get("/download")
def download_excel(
    customer_code: str = Query(..., description="Customer Code"),
    start_date: str = Query(..., description="Format: YYYYMMDD")
):
    # =========================
    # VALIDASI & KONVERSI TANGGAL
    # =========================
    start_dt = datetime.strptime(start_date, "%Y%m%d")
    # end_dt = start_dt + timedelta(days=1)

    # =========================
    # CONNECT REDSHIFT
    # =========================
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )

    # =========================
    # QUERY (AMAN - PARAMETERIZED)
    # =========================
    query = """
    SELECT 
        t1.connote__connote_code,
        t1.customer_code,
        t1.connote__connote_receiver_name,
        t1.connote__connote_receiver_address_detail,
        t1.connote__connote_state,
        t1.pod__timereceive,
        t2.pod__photo,
        t2.pod__signature
    FROM nipos.nipos t1
    JOIN nipos.nipos_pod_url t2
        ON t1.connote__connote_code = t2.connote__connote_code
    WHERE t1.customer_code = %s
        AND date (t1.connote__created_at)= %s
    """

    df = pd.read_sql(query, conn, params=(customer_code, start_date))
    conn.close()

    # =========================
    # BUAT EXCEL
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(df.columns.tolist())

    for _, row in df.iterrows():
        row_data = []
        for col in df.columns:
            if col in ["pod__photo", "pod__signature"]:
                row_data.append("")
            else:
                row_data.append(row[col])
        ws.append(row_data)

    # =========================
    # INSERT GAMBAR
    # =========================
    def insert_image_from_url(url, cell):
        if pd.isna(url):
            return
        try:
            response = requests.get(
                url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10
            )
            if response.status_code == 200:
                img_file = BytesIO(response.content)
                img = Image(img_file)

                max_size = 120
                ratio = min(max_size / img.width, max_size / img.height)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)

                ws.add_image(img, cell)
        except Exception:
            pass

    if "pod__photo" in df.columns and "pod__signature" in df.columns:

        photo_col = df.columns.get_loc("pod__photo") + 1
        sign_col = df.columns.get_loc("pod__signature") + 1

        for i in range(len(df)):
            excel_row = i + 2
            ws.row_dimensions[excel_row].height = 100

            photo_url = df.iloc[i]["pod__photo"]
            sign_url = df.iloc[i]["pod__signature"]

            photo_cell = ws.cell(row=excel_row, column=photo_col).coordinate
            sign_cell = ws.cell(row=excel_row, column=sign_col).coordinate

            insert_image_from_url(photo_url, photo_cell)
            insert_image_from_url(sign_url, sign_cell)

        ws.column_dimensions[
            ws.cell(row=1, column=photo_col).column_letter
        ].width = 20

        ws.column_dimensions[
            ws.cell(row=1, column=sign_col).column_letter
        ].width = 20

    # =========================
    # RETURN FILE
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"data_{customer_code}_{start_date}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
